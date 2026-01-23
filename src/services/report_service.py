from __future__ import annotations

import os
from pathlib import Path
from typing import Optional, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.models.enums import AttendanceStatus
from src.repositories.attendance_repo import AttendanceRepo
from src.repositories.class_repo import ClassRepo
from src.repositories.enrollment_repo import EnrollmentRepo
from src.repositories.session_repo import SessionRepo
from src.repositories.user_repo import UserRepo


class ReportService:
    """UC12 Summarize attendance; UC13 Export Excel."""

    def __init__(self) -> None:
        self._attendance_repo = AttendanceRepo()
        self._enrollment_repo = EnrollmentRepo()
        self._session_repo = SessionRepo()
        self._class_repo = ClassRepo()
        self._user_repo = UserRepo()

    def summarize(
        self, class_id: int, date_from: Optional[str] = None, date_to: Optional[str] = None
    ) -> list[dict[str, Any]]:
        """
        UC12: Summarize attendance - Tính tổng Present/Late/Absent/Excused per student.
        Trả về list summary với counts cho mỗi student.
        """
        # Lấy danh sách student enrolled trong class
        enrollments = self._enrollment_repo.list_by_filter(class_id=class_id)
        
        # Lấy attendance records cho class và date range
        sessions = self._session_repo.list_by_filter(
            class_id=class_id,
            date_from=date_from,
            date_to=date_to,
        )
        session_ids = [s.session_id for s in sessions]
        
        # Tính counts per student
        summaries = []
        for enrollment in enrollments:
            student_id = enrollment.student_id
            
            # Khởi tạo counts
            counts = {
                "student_id": student_id,
                "present": 0,
                "late": 0,
                "absent": 0,
                "excused": 0,
            }
            
            # Duyệt các session và đếm trạng thái
            for session_id in session_ids:
                record = self._attendance_repo.get_by_session_student(session_id, student_id)
                if record:
                    status = record.status
                    if status == AttendanceStatus.PRESENT.value:
                        counts["present"] += 1
                    elif status == AttendanceStatus.LATE.value:
                        counts["late"] += 1
                    elif status == AttendanceStatus.ABSENT.value:
                        counts["absent"] += 1
                    elif status == AttendanceStatus.EXCUSED.value:
                        counts["excused"] += 1
            
            counts["total"] = counts["present"] + counts["late"] + counts["absent"] + counts["excused"]
            summaries.append(counts)
        
        return summaries

    def export_excel(
        self, class_id: int, output_path: str, date_from: Optional[str] = None, date_to: Optional[str] = None
    ) -> str:
        """
        UC13: Export attendance report to Excel.
        Tạo 2 sheets: Summary (counts per student) + Detail (tất cả records).
        
        Args:
            class_id: Class ID
            output_path: Đường dẫn thư mục để lưu file
            date_from: Optional date filter (YYYY-MM-DD)
            date_to: Optional date filter (YYYY-MM-DD)
        
        Returns:
            Đường dẫn file Excel được tạo
        """
        # Ensure output directory exists
        os.makedirs(output_path, exist_ok=True)
        
        # Lấy class info
        class_info = self._class_repo.get_by_id(class_id)
        if not class_info:
            raise ValueError(f"Class {class_id} not found")
        
        # Tạo file name
        class_code = class_info.class_code
        file_name = f"Attendance_{class_code}.xlsx"
        file_path = os.path.join(output_path, file_name)
        
        # Tạo workbook
        wb = Workbook()
        
        # Sheet 1: Summary
        self._create_summary_sheet(wb, class_id, date_from, date_to)
        
        # Sheet 2: Detail
        self._create_detail_sheet(wb, class_id, date_from, date_to)
        
        # Xoá default sheet nếu có
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Lưu file
        wb.save(file_path)
        
        return file_path

    def _create_summary_sheet(
        self, wb: Workbook, class_id: int, date_from: Optional[str], date_to: Optional[str]
    ) -> None:
        """Tạo Summary sheet với counts Present/Late/Absent/Excused per student."""
        ws = wb.active
        ws.title = "Summary"
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Headers
        headers = ["Student ID", "Present", "Late", "Absent", "Excused", "Total"]
        ws.append(headers)
        
        # Format header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Lấy summary data
        summaries = self.summarize(class_id, date_from, date_to)
        
        # Ghi dữ liệu
        for summary in summaries:
            row = [
                summary["student_id"],
                summary["present"],
                summary["late"],
                summary["absent"],
                summary["excused"],
                summary["total"],
            ]
            ws.append(row)
        
        # Format data rows
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, 7):
                cell = ws.cell(row_idx, col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        for col in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 12

    def _create_detail_sheet(
        self, wb: Workbook, class_id: int, date_from: Optional[str], date_to: Optional[str]
    ) -> None:
        """Tạo Detail sheet với tất cả attendance records."""
        ws = wb.create_sheet("Detail")
        
        # Header style
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        
        # Headers
        headers = ["Session ID", "Session Date", "Student ID", "Status", "Check-in Time", "Note"]
        ws.append(headers)
        
        # Format header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Lấy sessions
        sessions = self._session_repo.list_by_filter(
            class_id=class_id,
            date_from=date_from,
            date_to=date_to,
        )
        
        # Lấy enrollment (student list)
        enrollments = self._enrollment_repo.list_by_filter(class_id=class_id)
        student_ids = [e.student_id for e in enrollments]
        
        # Ghi dữ liệu detail
        for session in sessions:
            for student_id in student_ids:
                record = self._attendance_repo.get_by_session_student(session.session_id, student_id)
                
                if record:
                    row = [
                        record.session_id,
                        session.session_date,
                        record.student_id,
                        record.status,
                        record.checkin_time or "",
                        record.note or "",
                    ]
                    ws.append(row)
        
        # Format data rows
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, 7):
                cell = ws.cell(row_idx, col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 25
